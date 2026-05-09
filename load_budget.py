#!/usr/bin/env python3
"""
load_budget.py — Build FY27_budget.sqlite from FY27 Budget .xlsx

Reads all 11 sheets across 3 schema families:
  - R-1 (21 cols, research summary)
  - P-1 (32 cols, procurement with quantity+amount pairs)
  - Service RDTE (17 cols × 9 sheets, PE/Project/narrative detail)

Usage:
    python load_budget.py
    python load_budget.py --xlsx "path/to/workbook.xlsx" --db "output.sqlite"
"""

import argparse
import math
import os
import sqlite3
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(SCRIPT_DIR, "FY27 Budget .xlsx")
DEFAULT_DB = os.path.join(SCRIPT_DIR, "FY27_budget.sqlite")

# ---------------------------------------------------------------------------
# Schema definitions
# ---------------------------------------------------------------------------

R1_SHEET = "FY27 R-1"
P1_SHEET = "FY27 P-1"
SERVICE_RDTE_SHEETS = [
    "FY27 USSF RDTE",
    "FY27 USAF RDTE",
    "FY27 USA RDTE",
    "FY27 USN RDTE",
    "FY27 MDA RDTE",
    "FY27 OSW RDTE",
    "FY27 TJS RDTE",
    "FY27 DARPA RDTE",
    "FY27 SOCOM RDTE",
]

CREATE_R1 = """
CREATE TABLE IF NOT EXISTS r1_research (
    id                    INTEGER PRIMARY KEY AUTOINCREMENT,
    account               TEXT,
    account_title         TEXT,
    organization          TEXT,
    budget_activity       TEXT,
    budget_activity_title TEXT,
    line_number           INTEGER,
    pe_bli                TEXT,
    pe_bli_title          TEXT,
    include_in_toa        TEXT,
    fy2025_actuals        REAL,
    fy2025_reconciliation REAL,
    fy2025_total          REAL,
    fy2026_disc_enacted   REAL,
    fy2026_pl119_spend    REAL,
    fy2026_total          REAL,
    fy2027_disc_request   REAL,
    fy2027_mandatory      REAL,
    fy2027_total          REAL,
    classification        TEXT,
    changes_notes         REAL,
    row_kind              TEXT DEFAULT 'line_item'
);
"""

CREATE_P1 = """
CREATE TABLE IF NOT EXISTS p1_procurement (
    id                    INTEGER PRIMARY KEY AUTOINCREMENT,
    account               TEXT,
    account_title         TEXT,
    organization          TEXT,
    budget_activity       TEXT,
    budget_activity_title TEXT,
    line_number           INTEGER,
    bsa                   TEXT,
    bsa_title             TEXT,
    budget_line_item      TEXT,
    bli_title             TEXT,
    cost_type             TEXT,
    cost_type_title       TEXT,
    add_non_add           TEXT,
    fy2025_qty            INTEGER,
    fy2025_amt            REAL,
    fy2025_recon_qty      INTEGER,
    fy2025_recon_amt      REAL,
    fy2025_total_qty      INTEGER,
    fy2025_total_amt      REAL,
    fy2026_enacted_qty    INTEGER,
    fy2026_enacted_amt    REAL,
    fy2026_pl119_qty      INTEGER,
    fy2026_pl119_amt      REAL,
    fy2026_total_qty      INTEGER,
    fy2026_total_amt      REAL,
    fy2027_request_qty    INTEGER,
    fy2027_request_amt    REAL,
    fy2027_mandatory_qty  INTEGER,
    fy2027_mandatory_amt  REAL,
    fy2027_total_qty      INTEGER,
    fy2027_total_amt      REAL,
    classification        TEXT
);
"""

CREATE_SERVICE_RDTE = """
CREATE TABLE IF NOT EXISTS service_rdte (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name      TEXT,
    line_number     TEXT,
    row_type        TEXT,
    pe_number       TEXT,
    project_number  TEXT,
    title           TEXT,
    budget_activity TEXT,
    description     TEXT,
    prior_years     REAL,
    fy2025          REAL,
    fy2026          REAL,
    fy2027_base     REAL,
    fy2027_ooc      REAL,
    fy2027_total    REAL,
    fy2028          REAL,
    fy2029          REAL,
    fy2030          REAL,
    fy2031          REAL
);
"""

# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

VIEWS = [
    (
        "v_fy27_request_by_pe",
        """
        CREATE VIEW IF NOT EXISTS v_fy27_request_by_pe AS
        SELECT
            pe_bli,
            pe_bli_title,
            account_title,
            budget_activity_title,
            fy2027_total,
            fy2026_total,
            CASE WHEN fy2026_total IS NOT NULL AND fy2026_total != 0
                 THEN fy2027_total - fy2026_total
                 ELSE NULL END AS delta_fy26_to_fy27,
            classification
        FROM r1_research
        WHERE include_in_toa = 'Y'
          AND row_kind = 'line_item'
        ORDER BY fy2027_total DESC NULLS LAST;
        """,
    ),
    (
        "v_fy27_delta",
        """
        CREATE VIEW IF NOT EXISTS v_fy27_delta AS
        SELECT
            pe_bli,
            pe_bli_title,
            account_title,
            fy2026_total,
            fy2027_total,
            changes_notes AS delta,
            CASE WHEN fy2026_total IS NOT NULL AND fy2026_total != 0
                 THEN ROUND((fy2027_total - fy2026_total) * 100.0 / fy2026_total, 1)
                 ELSE NULL END AS pct_change
        FROM r1_research
        WHERE include_in_toa = 'Y'
          AND row_kind = 'line_item'
          AND changes_notes IS NOT NULL
        ORDER BY changes_notes DESC;
        """,
    ),
    (
        "v_service_rdte_pe_summary",
        """
        CREATE VIEW IF NOT EXISTS v_service_rdte_pe_summary AS
        SELECT
            sheet_name,
            pe_number,
            title,
            budget_activity,
            fy2025,
            fy2026,
            fy2027_total,
            fy2028,
            fy2029,
            CASE WHEN fy2026 IS NOT NULL AND fy2026 != 0
                 THEN ROUND((fy2027_total - fy2026) * 100.0 / fy2026, 1)
                 ELSE NULL END AS pct_change_fy26_to_fy27
        FROM service_rdte
        WHERE row_type = 'PE'
        ORDER BY sheet_name, fy2027_total DESC NULLS LAST;
        """,
    ),
    (
        "v_procurement_by_program",
        """
        CREATE VIEW IF NOT EXISTS v_procurement_by_program AS
        SELECT
            account_title,
            budget_activity_title,
            bli_title,
            cost_type_title,
            fy2025_total_qty,
            fy2025_total_amt,
            fy2026_total_qty,
            fy2026_total_amt,
            fy2027_total_qty,
            fy2027_total_amt,
            classification
        FROM p1_procurement
        WHERE add_non_add = 'Add'
        ORDER BY fy2027_total_amt DESC NULLS LAST;
        """,
    ),
]

# ---------------------------------------------------------------------------
# Indexes for query performance
# ---------------------------------------------------------------------------

INDEXES = [
    "CREATE INDEX IF NOT EXISTS idx_r1_pe_bli ON r1_research(pe_bli);",
    "CREATE INDEX IF NOT EXISTS idx_r1_account ON r1_research(account);",
    "CREATE INDEX IF NOT EXISTS idx_r1_ba ON r1_research(budget_activity);",
    "CREATE INDEX IF NOT EXISTS idx_r1_row_kind ON r1_research(row_kind);",
    "CREATE INDEX IF NOT EXISTS idx_p1_bli ON p1_procurement(budget_line_item);",
    "CREATE INDEX IF NOT EXISTS idx_p1_account ON p1_procurement(account);",
    "CREATE INDEX IF NOT EXISTS idx_p1_add ON p1_procurement(add_non_add);",
    "CREATE INDEX IF NOT EXISTS idx_srdte_sheet ON service_rdte(sheet_name);",
    "CREATE INDEX IF NOT EXISTS idx_srdte_type ON service_rdte(row_type);",
    "CREATE INDEX IF NOT EXISTS idx_srdte_pe ON service_rdte(pe_number);",
    "CREATE INDEX IF NOT EXISTS idx_srdte_sheet_type ON service_rdte(sheet_name, row_type);",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def clean(val):
    """Convert Excel cell value to a Python value suitable for SQLite."""
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip()
        if stripped == "" or stripped.upper() == "NAN":
            return None
        return stripped
    if isinstance(val, float) and math.isnan(val):
        return None
    return val


def clean_int(val):
    """Convert to int if possible, else None."""
    v = clean(val)
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return v


def clean_float(val):
    """Convert to float if possible, else None."""
    v = clean(val)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return v


def classify_r1_row(line_number, pe_bli):
    """Determine row_kind for R-1 rows."""
    ln = clean(line_number)
    pe = clean(pe_bli)
    if ln is not None:
        try:
            if int(ln) == 999:
                return "classified_rollup"
        except (ValueError, TypeError):
            pass
    if pe is not None and str(pe).startswith("9999"):
        return "classified_rollup"
    return "line_item"


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def load_r1(wb, conn):
    """Load FY27 R-1 sheet into r1_research table."""
    ws = wb[R1_SHEET]
    cursor = conn.cursor()
    cursor.execute(CREATE_R1)

    rows_loaded = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # skip header
            continue
        if all(v is None for v in row):
            continue

        vals = list(row)
        # Pad or truncate to 21 columns
        while len(vals) < 21:
            vals.append(None)

        row_kind = classify_r1_row(vals[5], vals[6])

        cursor.execute(
            """INSERT INTO r1_research (
                account, account_title, organization, budget_activity,
                budget_activity_title, line_number, pe_bli, pe_bli_title,
                include_in_toa, fy2025_actuals, fy2025_reconciliation,
                fy2025_total, fy2026_disc_enacted, fy2026_pl119_spend,
                fy2026_total, fy2027_disc_request, fy2027_mandatory,
                fy2027_total, classification, changes_notes, row_kind
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                clean(vals[0]),
                clean(vals[1]),
                clean(vals[2]),
                clean(vals[3]),
                clean(vals[4]),
                clean_int(vals[5]),
                clean(vals[6]),
                clean(vals[7]),
                clean(vals[8]),
                clean_float(vals[9]),
                clean_float(vals[10]),
                clean_float(vals[11]),
                clean_float(vals[12]),
                clean_float(vals[13]),
                clean_float(vals[14]),
                clean_float(vals[15]),
                clean_float(vals[16]),
                clean_float(vals[17]),
                clean(vals[18]),
                clean_float(vals[20]),  # skip col 19 (unnamed)
                row_kind,
            ),
        )
        rows_loaded += 1

    conn.commit()
    print(f"  {R1_SHEET}: {rows_loaded} rows loaded")
    return rows_loaded


def load_p1(wb, conn):
    """Load FY27 P-1 sheet into p1_procurement table."""
    ws = wb[P1_SHEET]
    cursor = conn.cursor()
    cursor.execute(CREATE_P1)

    rows_loaded = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if all(v is None for v in row):
            continue

        vals = list(row)
        while len(vals) < 32:
            vals.append(None)

        cursor.execute(
            """INSERT INTO p1_procurement (
                account, account_title, organization, budget_activity,
                budget_activity_title, line_number, bsa, bsa_title,
                budget_line_item, bli_title, cost_type, cost_type_title,
                add_non_add,
                fy2025_qty, fy2025_amt,
                fy2025_recon_qty, fy2025_recon_amt,
                fy2025_total_qty, fy2025_total_amt,
                fy2026_enacted_qty, fy2026_enacted_amt,
                fy2026_pl119_qty, fy2026_pl119_amt,
                fy2026_total_qty, fy2026_total_amt,
                fy2027_request_qty, fy2027_request_amt,
                fy2027_mandatory_qty, fy2027_mandatory_amt,
                fy2027_total_qty, fy2027_total_amt,
                classification
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                clean(vals[0]),
                clean(vals[1]),
                clean(vals[2]),
                clean(vals[3]),
                clean(vals[4]),
                clean_int(vals[5]),
                clean(vals[6]),
                clean(vals[7]),
                clean(vals[8]),
                clean(vals[9]),
                clean(vals[10]),
                clean(vals[11]),
                clean(vals[12]),
                clean_int(vals[13]),
                clean_float(vals[14]),
                clean_int(vals[15]),
                clean_float(vals[16]),
                clean_int(vals[17]),
                clean_float(vals[18]),
                clean_int(vals[19]),
                clean_float(vals[20]),
                clean_int(vals[21]),
                clean_float(vals[22]),
                clean_int(vals[23]),
                clean_float(vals[24]),
                clean_int(vals[25]),
                clean_float(vals[26]),
                clean_int(vals[27]),
                clean_float(vals[28]),
                clean_int(vals[29]),
                clean_float(vals[30]),
                clean(vals[31]),
            ),
        )
        rows_loaded += 1

    conn.commit()
    print(f"  {P1_SHEET}: {rows_loaded} rows loaded")
    return rows_loaded


def load_service_rdte(wb, conn):
    """Load all 9 Service RDTE sheets into service_rdte table."""
    cursor = conn.cursor()
    cursor.execute(CREATE_SERVICE_RDTE)

    total = 0
    for sheet_name in SERVICE_RDTE_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        rows_loaded = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if all(v is None for v in row):
                continue

            vals = list(row)
            while len(vals) < 17:
                vals.append(None)

            cursor.execute(
                """INSERT INTO service_rdte (
                    sheet_name, line_number, row_type, pe_number,
                    project_number, title, budget_activity, description,
                    prior_years, fy2025, fy2026, fy2027_base, fy2027_ooc,
                    fy2027_total, fy2028, fy2029, fy2030, fy2031
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    sheet_name,
                    clean(vals[0]),
                    clean(vals[1]),
                    clean(vals[2]),
                    clean(vals[3]),
                    clean(vals[4]),
                    clean(vals[5]),
                    clean(vals[6]),
                    clean_float(vals[7]),
                    clean_float(vals[8]),
                    clean_float(vals[9]),
                    clean_float(vals[10]),
                    clean_float(vals[11]),
                    clean_float(vals[12]),
                    clean_float(vals[13]),
                    clean_float(vals[14]),
                    clean_float(vals[15]),
                    clean_float(vals[16]),
                ),
            )
            rows_loaded += 1

        conn.commit()
        print(f"  {sheet_name}: {rows_loaded} rows loaded")
        total += rows_loaded

    return total


def create_views(conn):
    """Create analytical views."""
    cursor = conn.cursor()
    for name, sql in VIEWS:
        cursor.execute(f"DROP VIEW IF EXISTS {name};")
        cursor.execute(sql)
        print(f"  View: {name}")
    conn.commit()


def create_indexes(conn):
    """Create indexes for query performance."""
    cursor = conn.cursor()
    for sql in INDEXES:
        cursor.execute(sql)
    conn.commit()
    print(f"  {len(INDEXES)} indexes created")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Load FY27 Budget workbook into SQLite")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to Excel workbook")
    parser.add_argument("--db", default=DEFAULT_DB, help="Output SQLite path")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: Workbook not found: {args.xlsx}")
        sys.exit(1)

    # Remove old DB if exists
    if os.path.exists(args.db):
        os.remove(args.db)
        print(f"Removed existing DB: {args.db}")

    print(f"Loading workbook: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    conn = sqlite3.connect(args.db)

    print("\n--- Loading R-1 ---")
    r1_count = load_r1(wb, conn)

    print("\n--- Loading P-1 ---")
    p1_count = load_p1(wb, conn)

    print("\n--- Loading Service RDTE ---")
    rdte_count = load_service_rdte(wb, conn)

    print("\n--- Creating views ---")
    create_views(conn)

    print("\n--- Creating indexes ---")
    create_indexes(conn)

    # Summary
    total = r1_count + p1_count + rdte_count
    print(f"\n{'='*50}")
    print(f"DONE — {total} total rows across 3 tables")
    print(f"  r1_research:    {r1_count:>6} rows")
    print(f"  p1_procurement: {p1_count:>6} rows")
    print(f"  service_rdte:   {rdte_count:>6} rows")
    print(f"DB written to: {args.db}")
    print(f"DB size: {os.path.getsize(args.db) / 1024 / 1024:.1f} MB")

    wb.close()
    conn.close()


if __name__ == "__main__":
    main()
